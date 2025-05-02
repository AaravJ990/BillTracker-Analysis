import streamlit as st
import openpyxl
from copy import copy
from datetime import datetime
import os

EXCEL_FILE = "my_bills.xlsx"

# Currency list with country names and fixed INR rates
CURRENCY_RATES = {
    "INR (India)": 1,
    "EUR (Eurozone)": 92.5,
    "AED (UAE)": 22.5,
    "AZN (Azerbaijan)": 56,
    "SEK (Sweden)": 8.55,
    "DKK (Denmark)": 12.41,
    "USD (United States)": 83.2,
    "GBP (United Kingdom)": 104.7,
    "JPY (Japan)": 0.56,
    "VND (Vietnam)": 0.0034,
    "LKR (Sri Lanka)": 0.27
    # Add more if needed
}

# Convert dropdown selection to currency code
def extract_currency_code(selection):
    return selection.split(" ")[0]

# Load Excel safely
def load_workbook_safe(path):
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    else:
        st.error(f"Excel file '{path}' not found.")
        st.stop()

# Copy cell style
def copy_row_style(ws, from_row, to_row, col_count):
    for col in range(1, col_count + 1):
        source = ws.cell(row=from_row, column=col)
        target = ws.cell(row=to_row, column=col)
        target._style = copy(source._style)
        if source.has_style:
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.number_format = copy(source.number_format)
            target.protection = copy(source.protection)
            target.alignment = copy(source.alignment)

# UI
st.title("📥 Bill Entry System")

with st.form("bill_form"):
    bill_date = st.date_input("📅 Date of Bill", datetime.today())

    # Fixed category dropdown
    fixed_categories = ["Flight", "Hotel", "Package", "Ride", "Food"]
    category = st.selectbox("🛍️ Category", fixed_categories)

    place = st.text_input("📍 Place")

    # Currency dropdown with country
    currency_display = st.selectbox("💱 Currency", list(CURRENCY_RATES.keys()))
    currency = extract_currency_code(currency_display)

    amount_foreign = st.number_input("💵 Amount in foreign currency", min_value=0.0, format="%.2f")
    link = st.text_input("🔗 Link to bill")

    submitted = st.form_submit_button("Add Bill")

if submitted:
    rate = CURRENCY_RATES[currency_display]
    amount_inr = round(amount_foreign * rate, 2)

    try:
        wb = load_workbook_safe(EXCEL_FILE)
    except PermissionError:
        st.error("The Excel file is open. Please close it and try again.")
        st.stop()

    ws = wb.active
    next_row = ws.max_row + 1
    prev_row = next_row - 1
    copy_row_style(ws, prev_row, next_row, 7)

    ws.cell(row=next_row, column=2).value = bill_date
    ws.cell(row=next_row, column=3).value = category
    ws.cell(row=next_row, column=4).value = place
    ws.cell(row=next_row, column=5).value = currency
    ws.cell(row=next_row, column=6).value = amount_foreign
    ws.cell(row=next_row, column=7).value = amount_inr
    ws.cell(row=next_row, column=8).value = link

    wb.save(EXCEL_FILE)

    st.success(f"✅ Bill added! {amount_foreign} {currency} = ₹{amount_inr}")
